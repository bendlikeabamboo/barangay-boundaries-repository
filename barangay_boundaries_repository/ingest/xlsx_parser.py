from __future__ import annotations

import re
from dataclasses import dataclass

import pandas as pd
from openpyxl import load_workbook

_CHANGE_TYPE_NORMALIZATION: dict[str, str] = {
    "change of code": "code_change",
    "city reverted as municipality": "reclassification",
    "converted from cc to huc": "reclassification",
    "correction of name": "renaming",
    "correction \nof name": "renaming",
    "deletion of barangay": "deletion",
    "deletion of municipality": "deletion",
    "division of barangay": "split",
    "division of province": "split",
    "fragmentation of a region": "split",
    "merged barangays": "merger",
    "merging of barangays": "merger",
    "merging of barangay": "merger",
    "municipality reverted as city": "reclassification",
    "new city": "creation",
    "new municipality": "creation",
    "new province": "creation",
    "new region": "creation",
    "new barangay": "creation",
    "newly created barangay": "creation",
    "newly created municipality": "creation",
    "newly created province": "creation",
    "newly created region": "creation",
    "newly created baragay": "creation",
    "re-enlisted": "reenlistment",
    "re-enlisted and correction \nof name": "reenlistment",
    "renaming": "renaming",
    "renaming of barangays": "renaming",
    "renaming of municipality": "renaming",
    "renaming of province": "renaming",
    "renaming of city": "renaming",
    "splitting of barangay": "split",
    "transfer of barangay": "transfer",
    "transfer of city": "transfer",
    "transfer of province": "transfer",
    "transfer of municipality": "transfer",
    "transferred municipality": "transfer",
    "transferred municipality and its barangay": "transfer",
    "newly created city": "creation",
    "deleted barangay": "deletion",
}


def _normalize_change_type(raw: str) -> str:
    key = raw.strip().lower()
    return _CHANGE_TYPE_NORMALIZATION.get(key, "unknown")


@dataclass
class PsgcRow:
    code: str
    name: str
    correspondence_code: str
    geographic_level: str
    old_name: str | None
    city_class: str | None
    income_class: str | None
    urban_rural: str | None
    population: int | None
    status: str | None


@dataclass
class ChangeEntry:
    entity_name: str
    unit_type_raw: str
    unit_type_normalized: str
    new_code: str
    mother_unit: str | None
    old_code: str | None
    description: str | None
    remarks: str | None
    section_date: str | None


@dataclass
class PsgcDatafile:
    snapshot_date: str
    rows: list[PsgcRow]
    notes: list[str]

    def by_level(self, level: str) -> list[PsgcRow]:
        return [r for r in self.rows if r.geographic_level == level]


@dataclass
class ChangeLog:
    snapshot_date: str
    entries: list[ChangeEntry]
    historical_entries: list[ChangeEntry]


def parse_datafile(path) -> PsgcDatafile:
    wb = load_workbook(path, read_only=True, data_only=True)

    snapshot_date = "unknown"
    if hasattr(path, "parent"):
        snapshot_date = path.parent.name

    df_psgc = pd.read_excel(path, sheet_name="PSGC", header=None)
    headers = df_psgc.iloc[0].tolist()
    df = df_psgc.iloc[1:].reset_index(drop=True)
    df.columns = range(len(headers))

    code_col = 0
    name_col = 1
    corr_col = 2
    level_col = 3
    old_name_col = 4
    city_class_col = 5
    income_col = 6
    urban_col = 7
    pop_col = 8
    status_col = 10

    rows: list[PsgcRow] = []
    for _, row in df.iterrows():
        code = str(row[code_col]).strip() if pd.notna(row[code_col]) else ""
        if not code or not re.fullmatch(r"\d{10}", code):
            continue
        name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
        corr_code = str(row[corr_col]).strip() if pd.notna(row[corr_col]) else ""
        level = str(row[level_col]).strip() if pd.notna(row[level_col]) else ""
        old_name = (
            str(row[old_name_col]).strip() if pd.notna(row[old_name_col]) else None
        )
        city_class = (
            str(row[city_class_col]).strip() if pd.notna(row[city_class_col]) else None
        )
        income_class = (
            str(row[income_col]).strip() if pd.notna(row[income_col]) else None
        )
        urban_rural = str(row[urban_col]).strip() if pd.notna(row[urban_col]) else None
        pop_val = row[pop_col]
        try:
            population = int(pop_val) if pd.notna(pop_val) else None
        except (ValueError, TypeError):
            population = None
        status = str(row[status_col]).strip() if pd.notna(row[status_col]) else None
        if status == "" or status == "nan":
            status = None

        rows.append(
            PsgcRow(
                code=code,
                name=name,
                correspondence_code=corr_code,
                geographic_level=level,
                old_name=old_name,
                city_class=city_class,
                income_class=income_class,
                urban_rural=urban_rural,
                population=population,
                status=status,
            )
        )

    notes: list[str] = []
    if "Notes" in wb.sheetnames:
        df_notes = pd.read_excel(path, sheet_name="Notes", header=None)
        for _, row in df_notes.iterrows():
            text = " ".join(str(v) for v in row if pd.notna(v) and str(v).strip())
            if text.strip():
                notes.append(text.strip())

    wb.close()
    return PsgcDatafile(snapshot_date=snapshot_date, rows=rows, notes=notes)


_SECTION_DATE_RE = re.compile(
    r"(January|February|March|April|May|June|July|August|September|October|November|December)"
    r"\s*(to|-|–)\s*"
    r"(January|February|March|April|May|June|July|August|September|October|November|December)"
    r"\s+(\d{4})",
    re.IGNORECASE,
)


def _parse_change_sheet(path, sheet_name: str) -> list[ChangeEntry]:
    df = pd.read_excel(path, sheet_name=sheet_name, header=None)
    entries: list[ChangeEntry] = []
    current_section: str | None = None

    for idx, row in df.iterrows():
        row_vals = [str(v).strip() if pd.notna(v) else "" for v in row]

        combined = " ".join(v for v in row_vals if v)
        if not combined:
            continue

        section_match = _SECTION_DATE_RE.search(combined)
        if section_match and row_vals[1] == "" and row_vals[2] == "":
            current_section = section_match.group(0)
            continue

        if row_vals[1].lower() in ("unit type", ""):
            continue

        if row_vals[1] == "" and row_vals[0] == "":
            continue

        if (
            row_vals[0] == "REPUBLIC OF THE PHILIPPINES"
            or row_vals[0] == "PHILIPPINE STATISTICS AUTHORITY"
        ):
            continue

        if not row_vals[0] and not row_vals[1]:
            continue

        entity_name = row_vals[0] if row_vals[0] else None
        unit_type_raw = row_vals[1] if row_vals[1] else None
        new_code = row_vals[2] if row_vals[2] else None
        mother_unit = row_vals[3] if row_vals[3] else None
        old_code = row_vals[4] if row_vals[4] else None
        description = row_vals[5] if row_vals[5] else None
        remarks = row_vals[6] if row_vals[6] else None

        if not entity_name and not unit_type_raw:
            if entries:
                last = entries[-1]
                if old_code and not last.old_code:
                    entries[-1] = ChangeEntry(
                        entity_name=last.entity_name,
                        unit_type_raw=last.unit_type_raw,
                        unit_type_normalized=last.unit_type_normalized,
                        new_code=last.new_code,
                        mother_unit=mother_unit or last.mother_unit,
                        old_code=old_code,
                        description=description or last.description,
                        remarks=remarks or last.remarks,
                        section_date=last.section_date,
                    )
            continue

        unit_type_normalized = (
            _normalize_change_type(unit_type_raw) if unit_type_raw else "unknown"
        )

        entries.append(
            ChangeEntry(
                entity_name=entity_name or "",
                unit_type_raw=unit_type_raw or "",
                unit_type_normalized=unit_type_normalized,
                new_code=new_code,
                mother_unit=mother_unit,
                old_code=old_code,
                description=description,
                remarks=remarks,
                section_date=current_section,
            )
        )

    return entries


def parse_changes(path) -> ChangeLog:
    snapshot_date = "unknown"
    if hasattr(path, "parent"):
        snapshot_date = path.parent.name

    wb = load_workbook(path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    entries: list[ChangeEntry] = []
    historical_entries: list[ChangeEntry] = []

    for name in sheet_names:
        if re.search(r"2001", name, re.IGNORECASE):
            entries = _parse_change_sheet(path, name)
        elif re.search(r"1977", name, re.IGNORECASE):
            historical_entries = _parse_change_sheet(path, name)

    return ChangeLog(
        snapshot_date=snapshot_date,
        entries=entries,
        historical_entries=historical_entries,
    )
